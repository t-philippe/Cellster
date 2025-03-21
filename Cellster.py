#Imports
import subprocess
import sys
import matplotlib
matplotlib.use('agg')
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import random as rd
from sklearn.cluster import DBSCAN
from openpyxl import Workbook
from scipy.spatial import ConvexHull
import tkinter as tk
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import tkinter as tk
from tkinter import filedialog
from shapely.geometry import Point, Polygon
import math
from scipy.spatial import cKDTree
from PIL import Image, ImageTk
from tkinter import ttk


#Variables
position_X = None
position_Y = None
ID = None
nb_clusters = None
nb_points = None
points = None
labels = None
set_label = None
clusters = None
nb_points_clusters = None
surface = None
densite = None
colors = None
file_open_path = None
file_save_path = None
cluster_ID = None
ID_new = None
nb_points_new = None
points_new = None
polygones = None
distances_poly = None
label_new = None
file_open_path_new = None
distances_points = None
points_min_dist = None
dessin_en_cours = False
forme_courbe = None
image_open_path = None
forme_poly = None

#Fonctions

##Extraire les données d'une fichier excel
def open_excel_file():
    global file_open_path
    fp = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls"), ("All Files", "*.*")])
    if fp:
            file_open_path = fp

def extract_position(filename_dir, sheetname='Position', header=1, name_pos_X='Position X', name_pos_Y='Position Y',name_ID='ID'):
    global position_X
    global position_Y
    global ID
    global nb_clusters
    global nb_points
    global points
    
    try:
        df = pd.read_excel(filename_dir,sheet_name=sheetname,header=header)
        position_X = df[name_pos_X]
        position_Y = df[name_pos_Y]
        ID = df[name_ID]
        nb_points = len(position_X)
        points = np.array([[position_X[i], position_Y[i]] for i in range(nb_points)])
        label_extract_conf["text"] = "Données chargées"
    except FileNotFoundError:
        label_extract_conf["text"] ="Fichier non trouvé"
    except Exception as e:
        label_extract_conf["text"] = "Une erreur est survenue"


def open_excel_file_new():
    global file_open_path_new
    fp = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls"), ("All Files", "*.*")])
    if fp:
            file_open_path_new = fp
            
        
def extract_new_position(filename_dir, sheetname='Position', header=1, name_pos_X='Position X', name_pos_Y='Position Y',name_ID='ID'):
    global ID_new
    global nb_points_new
    global points_new
    
    try:
        df = pd.read_excel(filename_dir,sheet_name=sheetname,header=header)
        position_X_new = df[name_pos_X]
        position_Y_new = df[name_pos_Y]
        ID_new = df[name_ID]
        nb_points_new = len(position_X_new)
        points_new = np.array([[position_X_new[i], position_Y_new[i]] for i in range(nb_points_new)])
        label_extract_conf_new["text"] = "Données chargées"
    except FileNotFoundError:
        label_extract_conf_new["text"] = "Fichier non trouvé"
    except Exception as e:
        label_extract_conf_new["text"] ="Une erreur est survenue"


def open_image():
    global image_open_path
    fp = filedialog.askopenfilename()
    if fp:
        image_open_path = fp

def draw_im(im_dir,frame):
    im = np.array(Image.open(im_dir))
    m=im.max()
    recomp = np.zeros((len(im),len(im[0])), dtype=np.uint8)
    for i in range(len(im)):
        for j in range(len(im[0])):
            a = (255*((im[i][j]/m)**2)).astype(np.uint8)
            recomp[i][j] = a
    image = Image.fromarray(recomp)
    nouvelle_taille = (600, 600)
    image_pil_redimensionnee = image.resize(nouvelle_taille, Image.LANCZOS)
    image_tk = ImageTk.PhotoImage(image_pil_redimensionnee)
    canevas_im.image_tk = image_tk
    canevas_im.delete("all")
    canevas_im.create_image(0, 0, anchor="nw", image=image_tk)
    
    


##Enregistrer les données dans un excel
def save_excel_file():
    global cluster_ID
    global points
    global ID
    global labels
    global nb_points_clusters
    global densite
    global surface
    
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Position"
    sheet2 = wb.create_sheet(title="Clusters")
    comp = 3
    comp2 = 3
    sheet.cell(row = 1, column = 1, value = 'Position')
    sheet.cell(row = 2, column = 1, value = 'Position X')
    sheet.cell(row = 2, column = 2, value = 'Position Y')
    sheet.cell(row = 2, column = 3, value = 'ID')
    sheet.cell(row = 2, column = 4, value = 'N° Cluster')
    
    sheet2.cell(row = 1, column = 1, value = 'Clusters')
    sheet2.cell(row = 2, column = 1, value = 'N° Cluster')
    sheet2.cell(row = 2, column = 2, value = 'Nb de points')
    sheet2.cell(row = 2, column = 3, value = 'Surface')
    sheet2.cell(row = 2, column = 4, value = 'Densité')

    for i in range(nb_points):
        sheet.cell(row = comp, column=1, value = points[i][0])
        sheet.cell(row = comp, column=2, value = points[i][1])
        sheet.cell(row = comp, column=3, value = ID[i])
        sheet.cell(row = comp, column=4, value = labels[i])
        comp += 1
            
    for i in cluster_ID:
        if surface[i] != 0:
            sheet2.cell(row = comp2, column = 1, value = i)
            sheet2.cell(row = comp2, column = 2, value = nb_points_clusters[i])
            sheet2.cell(row = comp2, column = 3, value = surface[i])
            sheet2.cell(row = comp2, column = 4, value = densite[i])
            comp2 += 1

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        wb.save(file_path)


def save_excel_file_new():
    global points_new
    global ID_new
    global label_new
    global nb_points_new
    global distances_poly
    
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Position"
    sheet2 = wb.create_sheet(title="Clusters")
    comp = 3
    comp2 = 3
    sheet.cell(row = 1, column = 1, value = 'Position')
    sheet.cell(row = 2, column = 1, value = 'Position X')
    sheet.cell(row = 2, column = 2, value = 'Position Y')
    sheet.cell(row = 2, column = 3, value = 'ID')
    sheet.cell(row = 2, column = 4, value = 'Distance au cluster')
    sheet.cell(row = 2, column = 5, value = 'N° Cluster')
    sheet.cell(row = 2, column = 6, value = 'Distance au point')
    sheet.cell(row = 2, column = 7, value = 'Coord point X')
    sheet.cell(row = 2, column = 8, value = 'Coord point Y')
    
    sheet2.cell(row = 1, column = 1, value = 'Clusters')
    sheet2.cell(row = 2, column = 1, value = 'N° Cluster')
    sheet2.cell(row = 2, column = 2, value = 'Nb de points')
    sheet2.cell(row = 2, column = 3, value = 'Densité')

    for i in range(nb_points_new):
        sheet.cell(row = comp, column = 1, value = points_new[i][0])
        sheet.cell(row = comp, column = 2, value = points_new[i][1])
        sheet.cell(row = comp, column = 3, value = ID_new[i])
        sheet.cell(row = comp, column = 4, value = distances_poly[i])
        sheet.cell(row = comp, column = 5, value = label_new[i])
        sheet.cell(row = comp, column = 6, value = distances_points[i])
        sheet.cell(row = comp, column = 7, value = points_min_dist[i][0])
        sheet.cell(row = comp, column = 8, value = points_min_dist[i][1])
        comp += 1
            
    for i in cluster_ID:
        if surface [i] != 0:
            res = sum([1 for j in range(nb_points_new) if label_new[j]==i])
            sheet2.cell(row = comp2, column = 1, value = i)
            sheet2.cell(row = comp2, column = 2, value = res)
            sheet2.cell(row = comp2, column = 3, value = res/surface[i])
            comp2 += 1

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        wb.save(file_path)
        

##Générer une couleur RGB aléatoire
def random_color_generator():
    r = rd.randint(0,255)/255
    g = rd.randint(0,255)/255
    b = rd.randint(0,255)/255
    return (r,g,b)


##Calcule de l'enveloppe concave
### From github.com/jsmolka/hull
def distance(p1, p2):
    """
    Calculates the distance between two points.

    :param p1, p2: points
    :return: distance between points
    """
    p1x, p1y = p1
    p2x, p2y = p2

    return math.sqrt((p1x - p2x) ** 2 + (p2y - p2y) ** 2)


def knn(points, p, k):
    """
    Calculates the k nearest neighbours of a point.

    :param points: list of points
    :param p: reference point
    :param k: amount of neighbours
    :return: list of k neighbours
    """
    return sorted(points, key=lambda x: distance(p, x))[:k]


def intersects(p1, p2, p3, p4):
    """
    Checks if the lines [p1, p2] and [p3, p4] intersect.

    :param p1, p2: line
    :param p3, p4: line
    :return: lines intersect
    """
    p0x, p0y = p1
    p1x, p1y = p2
    p2x, p2y = p3
    p3x, p3y = p4

    s10x = p1x - p0x
    s10y = p1y - p0y
    s32x = p3x - p2x
    s32y = p3y - p2y

    denom = s10x * s32y - s32x * s10y
    if denom == 0:
        return False

    denom_positive = denom > 0
    s02x = p0x - p2x
    s02y = p0y - p2y
    s_numer = s10x * s02y - s10y * s02x
    if (s_numer < 0) == denom_positive:
        return False

    t_numer = s32x * s02y - s32y * s02x
    if (t_numer < 0) == denom_positive:
        return False

    if (s_numer > denom) == denom_positive or (t_numer > denom) == denom_positive:
        return False

    t = t_numer / denom
    x = p0x + (t * s10x)
    y = p0y + (t * s10y)

    return (x, y) not in [p1, p2, p3, p4]


def angle(p1, p2, previous=0):
    """
    Calculates the angle between two points.

    :param p1, p2: points
    :param previous: previous angle
    :return: angle
    """
    p1x, p1y = p1
    p2x, p2y = p2

    return (math.atan2(p1y - p2y, p1x - p2x) - previous) % (math.pi * 2) - math.pi


def point_in_polygon(point, polygon):
    """
    Checks if a point is inside a polygon.

    :param point: point
    :param polygon: polygon
    :return: point is inside polygon
    """
    px, py = point

    size = len(polygon)
    for i in range(size):
        p1x, p1y = polygon[i]
        p2x, p2y = polygon[(i + 1) % size]
        if min(p1x, p2x) < px <= max(p1x, p2x):
            p = p1y - p2y
            q = p1x - p2x
            y = (px - p1x) * p / q + p1y
            if y < py:
                return True

    return False


def concave(points, k=3):
    """
    Calculates the concave hull for a list of points. Each point is a tuple
    containing the x- and y-coordinate. k defines the number of considered
    neighbours.

    :param points: list of points
    :param k: considered neighbours
    :return: concave hull
    """
    dataset = list(set(points))  # Remove duplicates
    if len(dataset) < 3:
        raise Exception("Dataset length cannot be smaller than 3")
    if len(dataset) == 3:
        return dataset  # Points are a polygon already
    
    k = min(max(k, 3), len(dataset) - 1)  # Make sure that k neighbours can be found

    first = current = min(dataset, key=lambda x: x[1])
    hull = [first]  # Initialize hull
    dataset.remove(first)  # Remove processed point
    previous_angle = 0

    while (current != first or len(hull) == 1) and len(dataset) > 0:
        if len(hull) == 3:
            dataset.append(first)  # Add first point again

        neighbours = knn(dataset, current, k)
        c_points = sorted(neighbours, key=lambda x: -angle(x, current, previous_angle))

        its = True
        i = -1
        while its and i < len(c_points) - 1:
            i += 1
            last_point = 1 if c_points[i] == first else 0
            j = 1
            its = False
            while not its and j < len(hull) - last_point:
                its = intersects(hull[-1], c_points[i], hull[-j - 1], hull[-j])
                j += 1

        if its:  # All points intersect, try again with higher a number of neighbours
            return concave(points, k + 1)

        previous_angle = angle(c_points[i], current)
        current = c_points[i]
        hull.append(current)  # Valid candidate was found
        dataset.remove(current)

    for point in dataset:
        if not point_in_polygon(point, hull):
            return concave(points, k + 1)

    return hull


##Construire les clusters
def clusterisation(distance, min_points):
    global points
    global labels
    global set_label
    global nb_clusters
    global colors
    global cluster_ID

    clustering = DBSCAN(eps = distance, min_samples = min_points).fit(points)
    labels = clustering.labels_
    set_label = set(labels)
    cluster_ID = [i for i in set_label if i!=-1]
    nb_clusters = len(set_label) - 1
    label_nb_clusters_1["text"] = "Nb clusters:"
    label_nb_clusters["text"] = "{}".format(nb_clusters)

    col=[]
    for i in range(nb_clusters):
        col.append(random_color_generator())
    colors=col

##Associer les numéros de chaque clusters
def labelisation():
    global clusters
    global clusters_ID
    
    res = []
    clus_ID = []
    for i in range(nb_clusters):
        res.append([points[j] for j in range(nb_points) if labels[j]==i])
        clus_ID.append([ID[j] for j in range(nb_points) if labels[j]==i])
    clusters = res
    cluster_ID = clus_ID
    

##Décrire les clusters
def description_clusters():
    global nb_points_clusters
    global surface
    global densite
    global polygones
    global points
    
    poly=[]
    surfa=[]
    densi=[]
    nbp=[]
    for i in range(nb_clusters):
        nb_points_clus=len(clusters[i])
        if nb_points_clus > 2:
            liste = [(poi[0],poi[1]) for poi in clusters[i]]
            hull = concave(liste,nb_points_clus//3)
            surf = Polygon(hull).area
            
            poly.append(Polygon(hull))
            nbp.append(nb_points_clus)
            surfa.append(surf)
            densi.append(nb_points_clus/surf)
        else:
            poly.append(Polygon(None))
            nbp.append(nb_points_clus)
            surfa.append(0)
            densi.append(0)
            

    nb_points_clusters = nbp
    surface = surfa
    densite = densi
    polygones = poly

##Afficher les informations d'un cluster
def get_info(num_cluster):

    if num_cluster in set_label:
        label_nbp["text"] = "Nb de points: {}".format(nb_points_clusters[num_cluster])
        label_sur["text"] = "Surface: {:.3f}".format(surface[num_cluster])
        label_den["text"] = "Densité: {:.6f}".format(densite[num_cluster])
    else:
        label_nbp["text"] = 'Erreur cluster'
        label_sur["text"] = ""
        label_den["text"] = ""
        
        
##Supprimer un cluster
def delete_cluster(num_cluster):
    global nb_clusters
    global labels
    global set_label
    global nb_points
    global cluster_ID
    
    for num in num_cluster:
        if num in set_label:
            for i in range(nb_points):
                    if labels[i] == num:
                        labels[i] = -1
            set_label.remove(num)
            nb_clusters = nb_clusters - 1
            cluster_ID.remove(num)
            label_nb_clusters["text"] ="{}".format(nb_clusters)
            label_delete_noti["text"] = "Supprimé".format(num_cluster)

        else:
            label_delete_noti["text"] = "Erreur cluster"

##Fusionner deux clusters            
def fusion(num_1,num_2):
    global nb_clusters
    global clusters
    global nb_point_clusters
    global nb_points
    global labels
    global set_label
    global surface
    global polygones
    global cluster_ID
    
    if (num_1 in set_label) and (num_2 in set_label):
        clusters[num_1]=clusters[num_1]+clusters[num_2]
        nb_clusters = nb_clusters - 1
        nb_points_clusters[num_1]=nb_points_clusters[num_1] + nb_points_clusters[num_2]
        for i in range(nb_points):
            if labels[i] == num_2:
                labels[i] = num_1
        set_label.remove(num_2)
        cluster_ID.remove(num_2)
        liste = [(poi[0],poi[1]) for poi in clusters[num_1]]
        hull = concave(liste,nb_points_clusters[num_1]//3)
        polygones[num_1] = Polygon(hull)
        set_changement=set()
        for i in range(nb_points):
            poi = Point(points[i][0],points[i][1])
            lab = labels[i]
            if labels[i]!= num_1 and polygones[num_1].distance(poi)==0:
                if lab !=-1:
                    set_changement.add(lab)
                    nb_points_clusters[lab] = nb_points_clusters[lab] - 1
                    for k in range(len(clusters[lab])-1,-1,-1):
                        if np.array_equal(np.array([position_X[i],position_Y[i]]),clusters[lab][k]):
                            clusters[lab].pop(k)
                labels[i] = num_1
                nb_points_clusters[num_1] = nb_points_clusters[num_1] + 1
                
        surface[num_1] = Polygon(hull).area
        densite[num_1] = nb_points_clusters[num_1] / surface[num_1]
        for c in set_changement:
            if nb_points_clusters[c] > 2:
                liste = [(poi[0],poi[1]) for poi in clusters[c]]
                hull = concave(liste,nb_points_clusters[c]//3)
                surf = Polygon(hull).area

                polygones[c] = Polygon(hull)
                surface[c] = surf
                densite[c] = nb_points_clusters[c]/surf
            else : 
                polygones[c] = Polygon(None)
                surface[c] = 0
                densite[c] = 0
                set_label.remove(c)
                cluster_ID.remove(c)
                nb_clusters = nb_clusters -1
        label_fusion_noti["text"] = "fusionné"
        label_nb_clusters["text"] = "{}".format(nb_clusters)
    else:
        label_fusion_noti["text"] = "Erreur cluster"       
        
        
##décrire la nouvelle liste           
def info_new_points():
    global nb_points_new
    global points_new
    global polygones
    global ID_new
    global distances_poly
    global label_new
    
    lab=[]
    dis=[]
    for i in range(nb_points_new):
        poi=Point(points_new[i][0],points_new[i][1])
        dis_ref = sys.maxsize
        label_ref = -1
        for i in set_label:
            if i != -1:
                dist = polygones[i].distance(poi)
                if dist < dis_ref:
                    dis_ref = dist
                    label_ref = i
        
        dis.append(dis_ref)
        if dis_ref<5:
            lab.append(label_ref)
        else:
            lab.append(-1)
    label_new = lab
    distances_poly = dis

##Calculer le point le plus proche de chaque point de la nouvelle liste
def contact():
    global points_new
    global points
    global nb_points
    global nb_points_new
    global distances_points
    global points_min_dist
    
    dist=[]
    points_min=[]
    tree = cKDTree(points)
    for i in range(nb_points_new):
        distance, index = tree.query(points_new[i])
        dist.append(distance)
        points_min.append(points[index])
        
    distances_points = dist
    points_min_dist = points_min
    

##Ajouter un cluster avec une forme
def ajout_manuel(frame):
    global fig, ax, forme_courbe
    global nb_points_clusters
    global clusters
    global set_label
    global cluster_ID
    global surface
    global polygones
    global nb_clusters
    global colors
    global labels

    if forme_courbe is None:

        for widget in frame.winfo_children():
            widget.pack_forget()

        fig, ax = plt.subplots()
        ax.scatter([position_X[i] for i in range(nb_points) if labels[i]==-1 ],[position_Y[i] for i in range(nb_points) if labels[i]==-1], s = 0.01, color = 'black')
        ax.set_title('Ajouter un cluster')

        for j in set_label:
            if j !=-1 and nb_points_clusters[j] > 2:
                ax.scatter([position_X[i] for i in range(nb_points) if labels[i]==j ],[position_Y[i] for i in range(nb_points) if labels[i]==j],label='cluster {}'.format(j),s=0.1,color=colors[j])
                x_center = sum(poi[0] for poi in clusters[j]) / len(clusters[j])
                y_center = sum(poi[1] for poi in clusters[j]) / len(clusters[j])
                ax.text(x_center, y_center, str(j),ha='center',va='center',fontsize=9)

        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.grid(row=0, column=0, sticky="nsew")
        canvas_widget.config(width=600, height=600)
        
        ###canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        canvas.get_tk_widget().bind("<MouseWheel>", zoom)
        canvas.get_tk_widget().bind("<Button-1>", debut_dessin)
        canvas.get_tk_widget().bind("<B1-Motion>", dessiner_forme)
        canvas.get_tk_widget().bind("<ButtonRelease-1>", fin_dessin)
        bouton_add_cluster["text"] = "Ajout"
        
    else:
        new_lab = len(clusters)
        polygon = Polygon(forme_courbe)
        nb = 0
        new_clus = []
        set_changement=set()
        for i in range(nb_points):
            poi = Point(points[i][0],points[i][1])
            lab = labels[i]
            if polygon.distance(poi)==0:
                new_clus.append([position_X[i],position_Y[i]])
                if lab != -1:
                    set_changement.add(lab)
                    clus = clusters[lab]
                    for k in range(len(clus)-1,-1,-1):
                        if np.array_equal(np.array([position_X[i],position_Y[i]]),clus[k]):
                            clusters[lab].pop(k)
                            nb_points_clusters[lab] -= 1
                labels[i] = new_lab
                nb += 1
                
        clusters += [new_clus]
        colors.append(random_color_generator())
        set_label.add(new_lab)
        nb_clusters += 1
        cluster_ID += [new_lab]
        nb_points_clusters.append(nb)
        liste = [(poi[0],poi[1]) for poi in new_clus]
        hull = concave(liste,nb//3)
        polygones.append(Polygon(hull))
        surface.append(Polygon(hull).area)
        densite.append(nb / Polygon(hull).area)
        for c in set_changement:
            if nb_points_clusters[c] > 2:
                liste = [(poi[0],poi[1]) for poi in clusters[c]]
                hull = concave(liste,nb_points_clusters[c]//3)
                surf = Polygon(hull).area

                polygones[c] = Polygon(hull)
                surface[c] = surf
                densite[c] = nb_points_clusters[c]/surf
            else : 
                polygones[c] = Polygon(None)
                surface[c] = 0
                densite[c] = 0
                set_label.remove(c)
                cluster_ID.remove(c)
                nb_clusters = nb_clusters -1
        label_nb_clusters["text"] = "{}".format(nb_clusters)
        bouton_add_cluster["text"] = "Ajout dessin"
        plot_clusters(frame)
        forme_courbe = None
        
def ajout_manuel_poly(frame):
    global fig, ax, forme_poly
    global nb_points_clusters
    global clusters
    global set_label
    global cluster_ID
    global surface
    global polygones
    global nb_clusters
    global colors
    global labels

    if forme_poly is None:

        for widget in frame.winfo_children():
            widget.pack_forget()
        forme_poly = []
        fig, ax = plt.subplots()
        ax.scatter([position_X[i] for i in range(nb_points) if labels[i]==-1 ],[position_Y[i] for i in range(nb_points) if labels[i]==-1], s = 0.01, color = 'black')
        ax.set_title('Ajouter un cluster')

        for j in set_label:
            if j !=-1 and nb_points_clusters[j] > 2:
                ax.scatter([position_X[i] for i in range(nb_points) if labels[i]==j ],[position_Y[i] for i in range(nb_points) if labels[i]==j],label='cluster {}'.format(j),s=0.1,color=colors[j])
                x_center = sum(poi[0] for poi in clusters[j]) / len(clusters[j])
                y_center = sum(poi[1] for poi in clusters[j]) / len(clusters[j])
                ax.text(x_center, y_center, str(j),ha='center',va='center',fontsize=9)

        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.grid(row=0, column=0, sticky="nsew")
        canvas_widget.config(width=600, height=600)
        ###canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        canvas.get_tk_widget().bind("<MouseWheel>", zoom)
        canvas.get_tk_widget().bind("<Button-1>", dessiner_polygone)
        bouton_add_cluster_poly["text"] = "Ajout"
        
    else:
        new_lab = len(clusters)
        polygon = Polygon(forme_poly)
        nb = 0
        new_clus = []
        set_changement=set()
        for i in range(nb_points):
            poi = Point(points[i][0],points[i][1])
            lab = labels[i]
            if polygon.distance(poi)==0:
                new_clus.append([position_X[i],position_Y[i]])
                if lab != -1:
                    set_changement.add(lab)
                    clus = clusters[lab]
                    for k in range(len(clus)-1,-1,-1):
                        if np.array_equal(np.array([position_X[i],position_Y[i]]),clus[k]):
                            clusters[lab].pop(k)
                            nb_points_clusters[lab] -= 1
                labels[i] = new_lab
                nb += 1
                
        clusters += [new_clus]
        colors.append(random_color_generator())
        set_label.add(new_lab)
        nb_clusters += 1
        cluster_ID += [new_lab]
        nb_points_clusters.append(nb)
        liste = [(poi[0],poi[1]) for poi in new_clus]
        hull = concave(liste,nb//3)
        polygones.append(Polygon(hull))
        surface.append(Polygon(hull).area)
        densite.append(nb / Polygon(hull).area)
        for c in set_changement:
            if nb_points_clusters[c] > 2:
                liste = [(poi[0],poi[1]) for poi in clusters[c]]
                hull = concave(liste,nb_points_clusters[c]//3)
                surf = Polygon(hull).area

                polygones[c] = Polygon(hull)
                surface[c] = surf
                densite[c] = nb_points_clusters[c]/surf
            else : 
                polygones[c] = Polygon(None)
                surface[c] = 0
                densite[c] = 0
                set_label.remove(c)
                cluster_ID.remove(c)
                nb_clusters = nb_clusters -1
        label_nb_clusters["text"] = "{}".format(nb_clusters)
        bouton_add_cluster_poly["text"] = "Ajout polygone"
        plot_clusters(frame)
        forme_poly = None

        
#Plots 

##Représenter l'ensemble des points extraits
def plot_points(frame):
    global fig, ax
    
    for widget in frame.winfo_children():
        widget.pack_forget()
        
    fig, ax = plt.subplots()
    ax.scatter(position_X, position_Y, s = 0.02, color= 'black')
    ax.set_title('Représentation des points')
    
    canvas = FigureCanvasTkAgg(fig, master=frame)
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.grid(row=0, column=0, sticky="nsew")
    canvas_widget.config(width=600, height=600)
    ###canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    canvas.get_tk_widget().bind("<MouseWheel>", zoom)
    canvas.get_tk_widget().bind("<Button-1>", )
    canvas.get_tk_widget().bind("<B1-Motion>", depla)
    canvas.get_tk_widget().bind("<ButtonRelease-1>", fin_depla)

##Représenter l'ensemble des clusters
def plot_clusters(frame):
    global fig, ax
    
    for widget in frame.winfo_children():
        widget.pack_forget()
        
    fig, ax = plt.subplots()    
    ax.scatter([position_X[i] for i in range(nb_points) if labels[i]==-1 ],[position_Y[i] for i in range(nb_points) if labels[i]==-1], s = 0.01, color = 'black')
    ax.set_title('Représentation des clusters')

    for j in set_label:
        if j !=-1 and nb_points_clusters[j] > 2:
            color=random_color_generator()
            ax.scatter([position_X[i] for i in range(nb_points) if labels[i]==j ],[position_Y[i] for i in range(nb_points) if labels[i]==j],label='cluster {}'.format(j),s=0.1,color=colors[j])
            x_center = sum(poi[0] for poi in clusters[j]) / len(clusters[j])
            y_center = sum(poi[1] for poi in clusters[j]) / len(clusters[j])
            ax.text(x_center, y_center, str(j),ha='center',va='center',fontsize=9)
        
    canvas = FigureCanvasTkAgg(fig, master=frame)
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.grid(row=0, column=0, sticky="nsew")
    canvas_widget.config(width=600, height=600)
    ###canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    canvas.get_tk_widget().bind("<MouseWheel>", zoom)
    canvas.get_tk_widget().bind("<Button-1>", debut_depla)
    canvas.get_tk_widget().bind("<B1-Motion>", depla)
    canvas.get_tk_widget().bind("<ButtonRelease-1>", fin_depla)

    
##Représenter une liste de clusters
def plot_list_clusters(text_label,frame):
    global fig, ax
    
    for widget in frame.winfo_children():
        widget.pack_forget()
        
    fig, ax = plt.subplots()
    ax.set_title('Liste clusters')
    
    ax.scatter([position_X[i] for i in range(nb_points) if labels[i]==-1 ],[position_Y[i] for i in range(nb_points) if labels[i]==-1], s = 0.01, color = 'black')
    
    if "-" in text_label:
        start = int(text_label.split("-")[0])
        end = int(text_label.split("-")[1])
        list_labels = [i for i in range(start,end+1)]
    else:
        list_labels = [int(elem) for elem in text_label.split(",")]
    for j in set_label:
        if j in list_labels and nb_points_clusters[j] > 2:
            ax.scatter([position_X[i] for i in range(nb_points) if labels[i]==j ],[position_Y[i] for i in range(nb_points) if labels[i]==j],label='cluster {}'.format(j),s=0.5,color=colors[j])
            x_center = sum(poi[0] for poi in clusters[j]) / len(clusters[j])
            y_center = sum(poi[1] for poi in clusters[j]) / len(clusters[j])
            ax.text(x_center, y_center, str(j),ha='center',va='center',fontsize=9)
        else:
            ax.scatter([position_X[i] for i in range(nb_points) if labels[i]==j ],[position_Y[i] for i in range(nb_points) if labels[i]==j],label='cluster {}'.format(j),s=0.01,color='black')

    ax.set_xlim(0,max(position_X))
    ax.set_ylim(0,max(position_Y))
    
    canvas = FigureCanvasTkAgg(fig, master=frame)
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.grid(row=0, column=0, sticky="nsew")
    canvas_widget.config(width=600, height=600)
    ###canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    canvas.get_tk_widget().bind("<MouseWheel>", zoom)
    canvas.get_tk_widget().bind("<Button-1>", debut_depla)
    canvas.get_tk_widget().bind("<B1-Motion>", depla)
    canvas.get_tk_widget().bind("<ButtonRelease-1>", fin_depla)
    
    
##Représenter le bruit
def plot_noise(frame):
    global fig, ax
    
    for widget in frame.winfo_children():
        widget.pack_forget()
        
    fig, ax = plt.subplots()
    ax.scatter([position_X[i] for i in range(nb_points) if labels[i]==-1 ],[position_Y[i] for i in range(nb_points) if labels[i]==-1], s = 0.01, color = 'black')
    ax.set_title('Représentation du bruit')
    
    canvas = FigureCanvasTkAgg(fig, master=frame)
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.grid(row=0, column=0, sticky="nsew")
    canvas_widget.config(width=600, height=600)
    ###canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    canvas.get_tk_widget().bind("<MouseWheel>", zoom)
    canvas.get_tk_widget().bind("<Button-1>", debut_depla)
    canvas.get_tk_widget().bind("<B1-Motion>", depla)
    canvas.get_tk_widget().bind("<ButtonRelease-1>", fin_depla)

##Représenter les nouveaux points
def plot_new_point(frame,taille):
    global fig, ax
    
    if  not points_new is None:
        for widget in frame.winfo_children():
            widget.pack_forget()
     
        fig, ax = plt.subplots()
        ax.scatter([position_X[i] for i in range(nb_points)],[position_Y[i] for i in range(nb_points)], s = 0.01, color = 'black')

        for j in set_label:
            if j !=-1 and nb_points_clusters[j] > 2:
                x_list=[]
                y_list=[]
                for sommet in list(polygones[j].exterior.coords):
                    x,y = sommet
                    x_list.append(x)
                    y_list.append(y)
                ax.plot(x_list,y_list,linewidth = 0.5)

        ax.set_title('Représentation des points')
        position_X_new=[poi[0] for poi in points_new]
        position_Y_new=[poi[1] for poi in points_new]

        ax.scatter(position_X_new, position_Y_new, s = taille, color= 'red')
        label_check["text"] = ""

        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.grid(row=0, column=0, sticky="nsew")
        canvas_widget.config(width=600, height=600)
        ###canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        canvas.get_tk_widget().bind("<MouseWheel>", zoom)
        canvas.get_tk_widget().bind("<Button-1>", debut_depla)
        canvas.get_tk_widget().bind("<B1-Motion>", depla)
        canvas.get_tk_widget().bind("<ButtonRelease-1>", fin_depla)
    else:
        label_check["text"] = "Erreur rouges"

##Représenter les surfaces
def plot_polygones(frame):  
    global fig, ax
    
    for widget in frame.winfo_children():
        widget.pack_forget()
        
    fig, ax = plt.subplots()
    ax.scatter([position_X[i] for i in range(nb_points)],[position_Y[i] for i in range(nb_points)], s = 0.01, color = 'black')

    for j in set_label:
        if j !=-1 and nb_points_clusters[j] > 2:
            x_list=[]
            y_list=[]
            for sommet in list(polygones[j].exterior.coords):
                x,y = sommet
                x_list.append(x)
                y_list.append(y)
            ax.plot(x_list,y_list,linewidth = 1)
            x_center = sum(poi[0] for poi in clusters[j]) / len(clusters[j])
            y_center = sum(poi[1] for poi in clusters[j]) / len(clusters[j])
            ax.text(x_center, y_center, str(j),ha='center',va='center',fontsize=9)
    ax.set_title('Représentation des points')
   
    canvas = FigureCanvasTkAgg(fig, master=frame)
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.grid(row=0, column=0, sticky="nsew")
    canvas_widget.config(width=600, height=600)
    ###canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    canvas.get_tk_widget().bind("<MouseWheel>", zoom)
    canvas.get_tk_widget().bind("<Button-1>", debut_depla)
    canvas.get_tk_widget().bind("<B1-Motion>", depla)
    canvas.get_tk_widget().bind("<ButtonRelease-1>", fin_depla)


#Manipuler les plots

##Zoom
def zoom(event):
    global fig, ax

    if fig and ax:
        x, y = ax.transData.inverted().transform([event.x, event.y])
        y = ax.get_ylim()[1] - y

        zoom_factor = 1.1
        if event.delta > 0: 
            zoom_factor = 1 / zoom_factor 
        new_xlim = [x - (x - ax.get_xlim()[0]) * zoom_factor, x + (ax.get_xlim()[1] - x) * zoom_factor]
        new_ylim = [y - (y - ax.get_ylim()[0]) * zoom_factor, y + (ax.get_ylim()[1] - y) * zoom_factor]
        ax.set_xlim(new_xlim)
        ax.set_ylim(new_ylim)
        fig.canvas.draw()  
        
##Commencer le déplacement
def debut_depla(event):
    global last_x, last_y
    last_x, last_y = event.x, event.y

##Déplacement
def depla(event):
    global last_x, last_y

    if last_x is not None and last_y is not None:
        dx = event.x - last_x
        dy = event.y - last_y
        dy = - dy * 2
        dx = dx * 2
        ax.set_xlim(ax.get_xlim()[0] - dx, ax.get_xlim()[1] - dx)
        ax.set_ylim(ax.get_ylim()[0] - dy, ax.get_ylim()[1] - dy)
        fig.canvas.draw()
        last_x, last_y = event.x, event.y

##Finir le déplacement
def fin_depla(event):
    global last_x, last_y
    last_x, last_y = None, None


##Dessiner forme pour cluster
def debut_dessin(event):
    global dessin_en_cours, forme_courbe
    dessin_en_cours = True
    forme_courbe = [] 
    x, y = ax.transData.inverted().transform([event.x, event.y])
    y = ax.get_ylim()[1] + ax.get_ylim()[0] - y
    forme_courbe.append((x, y))

##Dessiner la forme
def dessiner_forme(event):
    global forme_courbe

    if dessin_en_cours:
        x, y = ax.transData.inverted().transform([event.x, event.y])
        y = ax.get_ylim()[1] + ax.get_ylim()[0] - y
        forme_courbe.append((x, y))
        ax.plot(*zip(*forme_courbe), linestyle='-', color='red', label='Forme courbe',linewidth=0.1)
        fig.canvas.draw()
    
##Finir le dessin
def fin_dessin(event):
    global dessin_en_cours
    dessin_en_cours = False
    
##Dessiner un polygone
def dessiner_polygone(event):
    global forme_poly
    x, y = ax.transData.inverted().transform([event.x, event.y])
    y = ax.get_ylim()[1] + ax.get_ylim()[0] - y

    forme_poly.append((x, y))
    ax.plot(*zip(*forme_poly), linestyle='-', color='red', label='Forme courbe',linewidth=0.3)
    fig.canvas.draw()
    
    
class PlaceholderEntry:
    def __init__(self, master, placeholder,row,column,width = 20):
        self.placeholder = placeholder
        self.entry = ttk.Entry(master, style="Placeholder.TEntry",width=width)
        self.entry.insert(0, self.placeholder)
        self.entry.bind("<FocusIn>", self.on_entry_click)
        self.entry.bind("<FocusOut>", self.on_focus_out)
        self.entry.grid(row=row,column=column)
        self.entry.configure(foreground='grey') 


    def on_entry_click(self, event):
        if self.entry.get() == self.placeholder:
            self.entry.delete(0, "end")
            self.entry.configure(foreground='grey')

    def on_focus_out(self, event):
        if self.entry.get() == "":
            self.entry.insert(0, self.placeholder)
            self.entry.configure(foreground='grey')
            
    def get(self):
        return self.entry.get()

    
# Interface

##Créer la fenêtre
fenetre = tk.Tk()
fenetre.title("Traitement de clusters")
fenetre.grid_rowconfigure(0, weight=2)
fenetre.grid_rowconfigure(1, weight=1)
fenetre.geometry("1300x1000")


##Configurer les deux frames de séparation tracer et boutons
frame_up = tk.Frame(fenetre)
frame_up.grid(row=0, column=0, sticky="nsew")  # Occupe la colonne 0 et 4 lignes

frame_down = tk.Frame(fenetre)
frame_down.grid(row=1, column=0, sticky="nsew")  # Occupe la colonne 1 et 4 lignes
        
frame_up.grid_columnconfigure(0,weight=1)
frame_up.grid_columnconfigure(1,weight=1)

canevas_im = tk.Canvas(frame_up,width=600, height=600)
canevas_im.grid(row=0, column=1, sticky="nsew")
canevas_draw = tk.Canvas(frame_up,width=600, height=600)
canevas_draw.grid(row=0, column=0, sticky="nsew")



bouton_extract = tk.Button(frame_down, text = "Sélectionner Bleus", command = lambda : (open_excel_file(),extract_position(file_open_path)),width=15, height=1)     
bouton_extract.grid(row=0,column=0)
bouton_image = tk.Button(frame_down, text = "Sélectionner Image", command = lambda : (open_image(),draw_im(image_open_path,frame_up)),width=15, height=1)
bouton_image.grid(row=1,column=0,pady=2)
bouton_extract_new = tk.Button(frame_down, text = "Sélectionner Rouges", command = lambda : (open_excel_file_new(),extract_new_position(file_open_path_new)),width=15, height=1)     
bouton_extract_new.grid(row=2,column=0)

label_extract_conf_new = tk.Label(frame_down, text="")
label_extract_conf_new.grid(row=2,column=1)

label_extract_conf = tk.Label(frame_down, text="")
label_extract_conf.grid(row=0,column=1)


##Bouton pour tracer les points
bouton_plot = tk.Button(frame_down, text="Tracer les points", command=lambda: plot_points(frame_up))
bouton_plot.grid(row=1,column = 1,padx=2)

entry_dist = tk.Entry(frame_down,width=10)
entry_dist.grid(row=1,column=3)
entry_dist.insert(0,90)
label_dist = tk.Label(frame_down, text="Distance",width=10)
label_dist.grid(row=1,column=2)

entry_min_point = tk.Entry(frame_down,width=10)
entry_min_point.grid(row=2,column=3)
entry_min_point.insert(0,50)
label_min_point = tk.Label(frame_down, text="Nb de points")
label_min_point.grid(row=2,column=2)

bouton_clus = tk.Button(frame_down, text="Clusteriser", command=lambda: (clusterisation(float(entry_dist.get()),int(entry_min_point.get())),labelisation(),description_clusters(),plot_clusters(frame_up)))
bouton_clus.grid(row=0,column=3)
label_clus = tk.Label(frame_down, text="Clusterisation",width=10)
label_clus.grid(row=0,column=2)


label_nb_clusters_1=tk.Label(frame_down,text="")
label_nb_clusters_1.grid(row=0,column=4)
label_nb_clusters = tk.Label(frame_down, text="")
label_nb_clusters.grid(row=1,column=4)

bouton_noise = tk.Button(frame_down, text="Tracer le bruit", command=lambda: plot_noise(frame_up),width=12, height=1)
bouton_noise.grid(row=2,column=5)
bouton_polygones = tk.Button(frame_down, text="Tracer polygones", command=lambda: plot_polygones(frame_up),width=12, height=1)
bouton_polygones.grid(row=1,column=5)
bouton_clusters = tk.Button(frame_down, text="Tracer clusters", command = lambda:plot_clusters(frame_up),width=12, height=1)
bouton_clusters.grid(row=0,column=5)

##Espace représentation de clusters
label_tracer=tk.Label(frame_down,text="Dessin clusters")
label_tracer.grid(row=0,column=6)
entry_list =  PlaceholderEntry(frame_down, "liste: 2,5 début-fin: 1-6",1,6)
bouton_list_clusters= tk.Button(frame_down, text="Tracer la liste", command = lambda : plot_list_clusters(entry_list.get(),frame_up))
bouton_list_clusters.grid(row=2,column=6)

##Espace manipulation des clusters
label_manipuler=tk.Label(frame_down,text="Manipulation")
label_manipuler.grid(row=0,column=7)
bouton_add_cluster = tk.Button(frame_down,text="Ajout dessin", command = lambda : ajout_manuel(frame_up),width=12)
bouton_add_cluster.grid(row=1 ,column=7)
bouton_add_cluster_poly = tk.Button(frame_down,text="Ajout polygone", command = lambda : ajout_manuel_poly(frame_up),width=12)
bouton_add_cluster_poly.grid(row=2 ,column=7)

entry_delete = tk.Entry(frame_down)
entry_delete = PlaceholderEntry(frame_down, "N° à supprimer",1,8,14)
bouton_delete = tk.Button(frame_down,text = "Supprimer", command= lambda: (delete_cluster([int(elem) for elem in entry_delete.get().split(",")]),plot_clusters(frame_up)))
bouton_delete.grid(row=0,column=8,)
label_delete_noti = tk.Label(frame_down,text="")
label_delete_noti.grid(row=2,column=8)

entry_fusion_1 = PlaceholderEntry(frame_down, "Cluster 1",1,9,10)
bouton_fusion = tk.Button(frame_down, text = "Fusionner", command = lambda : (fusion(int(entry_fusion_1.get()),int(entry_fusion_2.get())),plot_clusters(frame_up)))
bouton_fusion.grid(row=0,column=9)
label_fusion_noti = tk.Label(frame_down,text="")
label_fusion_noti.grid(row=2,column=10)
entry_fusion_2 = PlaceholderEntry(frame_down,"Cluser 2",2,9,10)

entry_infos = PlaceholderEntry(frame_down, "N° cluster",1,10,10)
bouton_infos = tk.Button(frame_down,text='Infos', command= lambda: (get_info(int(entry_infos.get())),plot_list_clusters(entry_infos.get(),frame_up)))
bouton_infos.grid(row=0,column=10)

label_nbp = tk.Label(frame_down, text="")
label_nbp.grid(row=0,column=11)
label_sur = tk.Label(frame_down, text="")
label_sur.grid(row=1,column=11)
label_den = tk.Label(frame_down, text="")
label_den.grid(row=2,column=11)

bouton_save = tk.Button(frame_down, text='Enregistrer clusters', command = lambda : save_excel_file())
bouton_save.grid(row=0,column=13)
bouton_save_new = tk.Button(frame_down, text='Enregistrer Rouges', command = lambda : (info_new_points(),contact(),save_excel_file_new()))
bouton_save_new.grid(row=1,column=13)

entry_plot_new = PlaceholderEntry(frame_down,"Taille points",1,12,12)
bouton_plot_new = tk.Button(frame_down, text="Tracer Rouges", command = lambda: plot_new_point(frame_up,float(entry_plot_new.get())))
bouton_plot_new.grid(row=0,column = 12)
label_check = tk.Label(frame_down,text="")
label_check.grid(row=2,column=12)

fenetre.mainloop()
